from __future__ import annotations

import re
import sys
import zipfile
from pathlib import Path

from openpyxl import load_workbook


BASE_DIR = Path("D:/Graduate_test/dataset")
XLS_ROOT = BASE_DIR / "xls"


def find_template_dir() -> Path:
    for p in (XLS_ROOT / "成品").rglob("互感器_第一至第八层级图谱数据"):
        if p.is_dir():
            return p
    raise FileNotFoundError("未找到模板目录")


def header(ws_path: Path):
    wb = load_workbook(ws_path, read_only=True, data_only=True)
    ws = wb[wb.sheetnames[0]]
    h = list(next(ws.iter_rows(min_row=1, max_row=1, values_only=True)))
    rows = list(ws.iter_rows(values_only=True))
    wb.close()
    return h, rows


def get_sheet_path(z: zipfile.ZipFile) -> str:
    wb_xml = z.read("xl/workbook.xml").decode("utf-8", errors="replace")
    rid = re.search(r'<sheet[^>]*r:id="([^"]+)"', wb_xml).group(1)
    rels = z.read("xl/_rels/workbook.xml.rels").decode("utf-8", errors="replace")
    target = re.search(rf'<Relationship Id="{re.escape(rid)}"[^>]*Target="([^"]+)"', rels).group(1)
    return "xl/" + target.lstrip("/")


def cellxfs_count(z: zipfile.ZipFile) -> str:
    styles = z.read("xl/styles.xml").decode("utf-8", errors="replace")
    m = re.search(r'<cellXfs[^>]*count="(\d+)"', styles)
    return m.group(1) if m else "NA"


def validate(target_dir: Path, template_dir: Path) -> None:
    files = {
        "nodes": ("*_nodes.xlsx", "shared"),
        "links": ("*_links.xlsx", "shared"),
        "groups": ("*_groups.xlsx", "inline"),
        "members": ("*_group_members.xlsx", "inline"),
    }

    print("target_dir:", target_dir)
    print("template_dir:", template_dir)

    # 1) 业务校验
    n_file = next(target_dir.glob(files["nodes"][0]))
    l_file = next(target_dir.glob(files["links"][0]))
    g_file = next(target_dir.glob(files["groups"][0]))
    m_file = next(target_dir.glob(files["members"][0]))

    n_h, n_rows = header(n_file)
    l_h, l_rows = header(l_file)
    g_h, g_rows = header(g_file)
    m_h, m_rows = header(m_file)

    exp_n = ["id", "name", "degree", "desc", "fontFamily", "fontSize", "image", "lineWidth", "r", "stroke", "textBorderWidth", "weight"]
    exp_l = ["id", "from", "fromNodeName", "lineWidth", "relation", "stroke", "textBorderWidth", "to", "toNodeName"]
    exp_g = ["id", "name", "desc", "member_count", "order", "show", "type", "itemStyle.fill", "itemStyle.markColor", "itemStyle.stroke"]
    exp_m = ["id", "name", "member_id", "member_name"]

    if n_h[: len(exp_n)] != exp_n:
        raise ValueError(f"nodes 表头不符: {n_h}")
    if l_h[: len(exp_l)] != exp_l:
        raise ValueError(f"links 表头不符: {l_h}")
    if g_h[: len(exp_g)] != exp_g:
        raise ValueError(f"groups 表头不符: {g_h}")
    if m_h[: len(exp_m)] != exp_m:
        raise ValueError(f"members 表头不符: {m_h}")

    node_ids = {str(r[0]) for r in n_rows[1:] if r and r[0] is not None}
    if len(node_ids) != len(n_rows) - 1:
        raise ValueError("nodes id 不唯一")

    for r in l_rows[1:]:
        if str(r[1]) not in node_ids or str(r[7]) not in node_ids:
            raise ValueError(f"links 引用断裂: {r}")

    group_ids = {str(r[0]) for r in g_rows[1:] if r and r[0] is not None}
    for r in m_rows[1:]:
        if str(r[0]) not in group_ids or str(r[2]) not in node_ids:
            raise ValueError(f"group_members 引用断裂: {r}")

    # 2) 壳校验（对比模板）
    for key, (pat, mode) in files.items():
        tf = next(template_dir.glob(pat))
        of = next(target_dir.glob(pat))
        with zipfile.ZipFile(tf) as zt, zipfile.ZipFile(of) as zo:
            if cellxfs_count(zt) != cellxfs_count(zo):
                raise ValueError(f"{key} cellXfs 不一致")
            if mode == "shared" and "xl/sharedStrings.xml" not in zo.namelist():
                raise ValueError(f"{key} 缺少 sharedStrings.xml")
            sheet = zo.read(get_sheet_path(zo)).decode("utf-8", errors="replace")
            if mode == "shared" and 't="inlineStr"' in sheet:
                raise ValueError(f"{key} 使用了 inlineStr，非壳兼容写法")

    print("business_check: OK")
    print("shell_check: OK")
    print(
        "counts:",
        f"nodes={len(n_rows)-1}",
        f"links={len(l_rows)-1}",
        f"groups={len(g_rows)-1}",
        f"members={len(m_rows)-1}",
    )


def main() -> None:
    target_dir = Path(sys.argv[1]) if len(sys.argv) > 1 else (XLS_ROOT / "generated_mutual_import_styled")
    template_dir = find_template_dir()
    validate(target_dir, template_dir)


if __name__ == "__main__":
    main()
