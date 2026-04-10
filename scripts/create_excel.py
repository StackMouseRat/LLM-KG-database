# -*- coding: utf-8 -*-
from openpyxl import load_workbook, Workbook
from openpyxl.styles import Font, PatternFill

# 读取节点模板
nodes_template_path = r'C:\Users\StackRat\Desktop\docker\dataset\xls\NRD Studio Excel模板文件\节点_nodes.xlsx'
links_template_path = r'C:\Users\StackRat\Desktop\docker\dataset\xls\NRD Studio Excel模板文件\关系_links.xlsx'

# 输出文件路径
nodes_output_path = r'C:\Users\StackRat\Desktop\docker\dataset\xls\节点.xlsx'
links_output_path = r'C:\Users\StackRat\Desktop\docker\dataset\xls\关系.xlsx'

# ==================== 读取模板结构 ====================
print("Reading templates...")

# 读取节点模板
wb_nodes_template = load_workbook(nodes_template_path)
ws_nodes_template = wb_nodes_template.active
nodes_columns = [cell.value for cell in ws_nodes_template[1]]
print(f"Nodes columns: {nodes_columns}")

# 读取关系模板
wb_links_template = load_workbook(links_template_path)
ws_links_template = wb_links_template.active
links_columns = [cell.value for cell in ws_links_template[1]]
print(f"Links columns: {links_columns}")

# ==================== 定义数据 ====================

# 节点数据
nodes_data = [
    # (节点名称, 描述, 图标, 权重, 组, 类型值, 类型, 形状, 填充颜色, 描边颜色)
    ("电气设备实体", "电气设备实体", "https://nrdstudio.cn/res/n.jpg", 1, 0, 1, None, None, None, None),
    ("故障类型", "故障类型", "https://nrdstudio.cn/res/n.jpg", 2, 0, 2, None, None, None, None),
    ("故障原因", "故障原因", "https://nrdstudio.cn/res/n.jpg", 3, 0, 3, None, None, None, None),
    ("故障现象", "故障现象", "https://nrdstudio.cn/res/n.jpg", 4, 0, 4, None, None, None, None),
    ("故障后果", "故障后果", "https://nrdstudio.cn/res/n.jpg", 5, 0, 5, None, None, None, None),
    ("应对措施", "应对措施", "https://nrdstudio.cn/res/n.jpg", 6, 0, 6, None, None, None, None),
    ("安全风险", "安全风险", "https://nrdstudio.cn/res/n.jpg", 7, 0, 7, None, None, None, None),
    ("应急资源", "应急资源", "https://nrdstudio.cn/res/n.jpg", 8, 0, 8, None, None, None, None),
]

# 关系数据
links_data = [
    # (源节点, 目标节点, 关系类型)
    ("电气设备实体", "故障类型", "发生"),
    ("故障类型", "故障原因", "起因于"),
    ("故障类型", "故障现象", "表现为"),
    ("故障类型", "故障后果", "导致"),
    ("故障类型", "应对措施", "处置"),
    ("故障原因", "应对措施", "针对性"),
    ("应对措施", "安全风险", "存在"),
    ("应对措施", "应急资源", "需要"),
]

# ==================== 创建节点Excel文件 ====================
print(f"\nCreating nodes file: {nodes_output_path}")
wb_nodes = Workbook()
ws_nodes = wb_nodes.active

# 写入表头
header_fill = PatternFill(start_color="CCE5FF", end_color="CCE5FF", fill_type="solid")
header_font = Font(bold=True)

for col_idx, col_name in enumerate(nodes_columns, 1):
    cell = ws_nodes.cell(row=1, column=col_idx)
    cell.value = col_name
    cell.fill = header_fill
    cell.font = header_font

# 写入数据
for row_idx, row_data in enumerate(nodes_data, 2):
    for col_idx, value in enumerate(row_data, 1):
        ws_nodes.cell(row=row_idx, column=col_idx, value=value)

wb_nodes.save(nodes_output_path)
print(f"Nodes file created with {len(nodes_data)} rows")

# ==================== 创建关系Excel文件 ====================
print(f"\nCreating links file: {links_output_path}")
wb_links = Workbook()
ws_links = wb_links.active

# 写入表头
for col_idx, col_name in enumerate(links_columns, 1):
    cell = ws_links.cell(row=1, column=col_idx)
    cell.value = col_name
    cell.fill = header_fill
    cell.font = header_font

# 写入数据
for row_idx, row_data in enumerate(links_data, 2):
    for col_idx, value in enumerate(row_data, 1):
        ws_links.cell(row=row_idx, column=col_idx, value=value)

wb_links.save(links_output_path)
print(f"Links file created with {len(links_data)} rows")

print("\n=== Files created successfully! ===")
