# -*- coding: utf-8 -*-
from openpyxl import load_workbook
import sys

# 设置输出编码为UTF-8
sys.stdout.reconfigure(encoding='utf-8')

# 验证节点文件
print("=== 节点.xlsx 内容 ===")
wb = load_workbook(r'C:\Users\StackRat\Desktop\docker\dataset\xls\节点.xlsx')
ws = wb.active
for i, row in enumerate(ws.iter_rows(values_only=True), 1):
    print(f"行 {i}: {row}")

print("\n=== 关系.xlsx 内容 ===")
wb = load_workbook(r'C:\Users\StackRat\Desktop\docker\dataset\xls\关系.xlsx')
ws = wb.active
for i, row in enumerate(ws.iter_rows(values_only=True), 1):
    print(f"行 {i}: {row}")

print("\n=== 文件创建成功! ===")
print("文件位置:")
print("- C:\\Users\\StackRat\\Desktop\\docker\\dataset\\xls\\节点.xlsx")
print("- C:\\Users\\StackRat\\Desktop\\docker\\dataset\\xls\\关系.xlsx")
