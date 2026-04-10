# -*- coding: utf-8 -*-
"""
处理输电线路故障数据并生成NRD Studio格式的Excel文件
"""

import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Font, Alignment
from copy import copy

# 定义从文档中提取的数据
# 电气设备
EQUIPMENT = '输电线路'

# 故障类型及详细信息
# 键为故障名称，值为包含原因、现象、后果、措施的字典
FAULTS = {
    '覆冰过荷载': {
        '原因': [
            '对输电线路覆冰的规律和微气象环境认识不足，线路路径选择不合理，抗冰设计不足',
            '输电线路的设计抗冰厚度低于实际覆冰值',
            '气候极其恶劣，存在薄弱环节，遭遇超过气象记录的恶劣气候条件'
        ],
        '现象': [
            '导线断裂',
            '地线断裂',
            '杆塔倒塔',
            '金具损坏',
            '短路跳闸',
            '烧伤甚至烧断导线',
            '导线从压接管内抽出',
            '外层铝股全断、钢芯抽出',
            '杆塔基础下沉、倾斜或爆裂'
        ],
        '后果': [
            '线路停电',
            '设备损坏',
            '供电中断'
        ],
        '措施': [
            '增加绝缘子串长度，提高绝缘子串的绝缘水平',
            '改变绝缘子的悬挂方式，使用倒V串结构',
            '采用大盘径绝缘子插花方式',
            '采用防冰涂料或防污闪涂料',
            '保持绝缘子和塔头上部的清洁',
            '增大杆塔覆冰、舞动的设防值',
            '重冰区域杆塔采用高强度钢',
            '采用高强度、大容量、自重轻的导线',
            '采用高强度，耐低温的新型复合绝缘子',
            '改进重冰区输电线路使用金具，增大强度、减轻质量'
        ]
    },
    '脱冰跳跃': {
        '原因': [
            '相邻档导线不均匀覆冰',
            '不同期脱冰'
        ],
        '现象': [
            '导线缩紧和断裂',
            '绝缘子损伤和破裂',
            '杆塔横担扭转和变形',
            '线间电气间隙减小',
            '导线放电烧伤'
        ],
        '后果': [
            '设备损坏',
            '线路跳闸'
        ],
        '措施': [
            '增加绝缘子串长度，提高绝缘子串的绝缘水平',
            '改变绝缘子的悬挂方式，使用倒V串结构',
            '采用高强度、大容量、自重轻的导线',
            '改进重冰区输电线路使用金具，增大强度、减轻质量'
        ]
    },
    '覆冰闪络': {
        '原因': [
            '春冬季时期持续电弧烧伤绝缘子',
            '绝缘子表面污秽'
        ],
        '现象': [
            '闪络过程中持续电弧烧伤绝缘子',
            '绝缘强度下降'
        ],
        '后果': [
            '线路跳闸',
            '设备损坏'
        ],
        '措施': [
            '增加绝缘子串长度，提高绝缘子串的绝缘水平',
            '改变绝缘子的悬挂方式，使用倒V串结构',
            '采用大盘径绝缘子插花方式',
            '采用防冰涂料或防污闪涂料',
            '保持绝缘子和塔头上部的清洁'
        ]
    },
    '覆冰舞动': {
        '原因': [
            '不均匀覆冰使导线产生自激振荡和舞动'
        ],
        '现象': [
            '金具损坏',
            '导线断股',
            '杆塔倾斜或倒塔'
        ],
        '后果': [
            '线路停电',
            '设备损坏'
        ],
        '措施': [
            '增大杆塔覆冰、舞动的设防值',
            '重冰区域杆塔采用高强度钢',
            '采用高强度、大容量、自重轻的导线',
            '缠绕扰流线',
            '安装空气动力阻尼器',
            '安装扭转阻尼器、失谐摆、抑扭环',
            '安装压重防舞装置',
            '安装相间间隔棒',
            '安装偏心重锤双摆防舞器',
            '线夹回转式间隔棒、双摆防舞器'
        ]
    },
    '舞动导致螺栓松脱': {
        '原因': [
            '杆塔连接处螺孔与螺栓存在连接间隙',
            '交变荷载造成预紧力下降',
            '投运时间短'
        ],
        '现象': [
            '螺栓松脱',
            '连接杆件相互滑移'
        ],
        '后果': [
            '金具损坏',
            '结构失稳'
        ],
        '措施': [
            '选择适合工程的防松螺帽',
            '对杆塔紧固件进行有计划的二次复紧'
        ]
    },
    '舞动导致杆塔损坏': {
        '原因': [
            '导线舞动荷载直接作用在杆塔上',
            '绝缘子摆动传递舞动荷载',
            '螺栓松动脱落'
        ],
        '现象': [
            '间隔棒夹头松动或脱落',
            '球头挂环磨坏断裂',
            '耐张塔损坏',
            '直线塔损坏',
            '横担扭曲变形及断裂'
        ],
        '后果': [
            '导线落地',
            '结构损坏'
        ],
        '措施': [
            '合理选用金具',
            '采用防松螺帽、双螺帽',
            '合理确定线间距离',
            '缩短耐张段'
        ]
    },
    '舞动导致跳闸': {
        '原因': [
            '导线舞动导致相间短路',
            '对杆塔放电',
            '对下方交叉线路或物体放电',
            '导线损坏',
            '杆塔损坏'
        ],
        '现象': [
            '相间闪络',
            '跳闸',
            '短路'
        ],
        '后果': [
            '线路停电',
            '影响供电可靠性'
        ],
        '措施': [
            '合理选用金具',
            '采用防松螺帽、双螺帽',
            '合理确定线间距离',
            '缩短耐张段',
            '缠绕扰流线',
            '安装空气动力阻尼器',
            '安装扭转阻尼器、失谐摆、抑扭环',
            '安装压重防舞装置',
            '安装相间间隔棒',
            '安装偏心重锤双摆防舞器'
        ]
    },
    '反击雷': {
        '原因': [
            '雷击于杆塔顶部或杆塔附近的避雷线'
        ],
        '现象': [
            '雷电过电压',
            '绝缘子闪络'
        ],
        '后果': [
            '线路跳闸',
            '设备损坏'
        ],
        '措施': [
            '减小避雷线保护角',
            '降低杆塔接地电阻',
            '安装线路避雷器',
            '架设耦合地线',
            '加装各种型式避雷针'
        ]
    },
    '绕击雷': {
        '原因': [
            '雷绕过避雷线击于导线'
        ],
        '现象': [
            '雷电过电压',
            '绝缘子闪络'
        ],
        '后果': [
            '线路跳闸',
            '设备损坏'
        ],
        '措施': [
            '减小避雷线保护角',
            '降低杆塔接地电阻',
            '安装线路避雷器',
            '架设耦合地线'
        ]
    },
    '雷击档距中央': {
        '原因': [
            '雷击于避雷线档距中央'
        ],
        '现象': [
            '雷电过电压',
            '绝缘子闪络'
        ],
        '后果': [
            '线路跳闸',
            '设备损坏'
        ],
        '措施': [
            '减小避雷线保护角',
            '降低杆塔接地电阻',
            '安装线路避雷器'
        ]
    },
    '异物短路': {
        '原因': [
            '高空抛物',
            '漂浮悬挂的气球',
            '风筝',
            '大风吹倒树木压在导线上',
            '广告布、遮阳布等漂浮物'
        ],
        '现象': [
            '相间或相地连接',
            '短接空气间隙',
            '线路闪络',
            '悬挂在导线和杆塔上'
        ],
        '后果': [
            '线路跳闸',
            '故障停运'
        ],
        '措施': [
            '大力整治清理线路通道',
            '加强线路巡视',
            '充分发动沿线群众护线',
            '加强护线宣传教育',
            '安装移动视频监控及预警装置'
        ]
    },
    '吊车碰线': {
        '原因': [
            '吊车在线路旁施工',
            '转臂或起吊移动过程中对导线距离过近'
        ],
        '现象': [
            '相地之间短路',
            '导线断线'
        ],
        '后果': [
            '线路跳闸',
            '设备损坏',
            '施工人员触电风险'
        ],
        '措施': [
            '落实各级设备主人制',
            '设置线路外部隐患专责人',
            '对重大外部隐患实行定点蹲守',
            '安装移动视频监控及预警装置',
            '加强护线宣传教育'
        ]
    },
    '违章施工': {
        '原因': [
            '放炮',
            '违章建房',
            '导电物体与导线之间的安全距离过近'
        ],
        '现象': [
            '导线损伤',
            '危及杆塔基础',
            '线路跳闸'
        ],
        '后果': [
            '线路跳闸',
            '杆塔倒塌风险',
            '人员触电风险'
        ],
        '措施': [
            '落实各级设备主人制',
            '设置线路外部隐患专责人',
            '对重大外部隐患实行定点蹲守',
            '合理利用低压保高压政策',
            '加强护线宣传教育'
        ]
    },
    '烟火短路': {
        '原因': [
            '放礼花弹等烟火',
            '烟雾使空气绝缘性能下降',
            '礼花弹弹出的锡纸彩带'
        ],
        '现象': [
            '相间或相地击穿',
            '闪络'
        ],
        '后果': [
            '线路跳闸'
        ],
        '措施': [
            '加强护线宣传教育',
            '发放宣传画册',
            '树立警示牌',
            '举办培训班'
        ]
    },
    '山火': {
        '原因': [
            '山林火灾',
            '导线周围空气电离',
            '烟雾携带大量导电物质'
        ],
        '现象': [
            '空气绝缘性能下降',
            '导线对地线闪络',
            '导线之间闪络',
            '导线对地闪络'
        ],
        '后果': [
            '线路跳闸',
            '设备损坏',
            '大面积停电'
        ],
        '措施': [
            '建立并固化外力破坏隐患流程',
            '完善群众护线网络',
            '加强护线宣传教育',
            '根据季节特点和山火发生规律加强防山火预测及预警'
        ]
    },
    '盗窃': {
        '原因': [
            '不法分子偷拆塔材',
            '偷割拉线',
            '偷拆UT线夹',
            '偷割导线电缆'
        ],
        '现象': [
            '塔材缺失',
            '拉线缺失',
            '倒杆塔',
            '断线'
        ],
        '后果': [
            '线路故障',
            '大面积停电',
            '重大经济损失'
        ],
        '措施': [
            '完善群众护线网络',
            '建立并完善的护线网络',
            '设置防外力破坏联防举报点',
            '加强与政府职能部门联系',
            '严厉打击电力设施偷盗行为'
        ]
    },
    '倒塔断线': {
        '原因': [
            '风荷载超过杆塔极限荷载',
            '导线极限荷载超过设计值'
        ],
        '现象': [
            '杆塔倒塔',
            '导地线断线'
        ],
        '后果': [
            '线路停电',
            '重大设备损坏'
        ],
        '措施': [
            '适当提高抵御台风的设计标准',
            '采取局部加强措施',
            '线路杆塔结构重要性系数取1.1',
            '明确防风害差异化改造原则'
        ]
    },
    '风偏闪络': {
        '原因': [
            '强风使导线风偏角过大',
            '暴雨降低空气间隙的放电电压',
            '设计裕度不足'
        ],
        '现象': [
            '闪络跳闸'
        ],
        '后果': [
            '线路停电',
            '影响供电可靠性'
        ],
        '措施': [
            '规范跳线安装和改造措施',
            '采用双绝缘子串加装支撑管改造',
            '跳线引线整治措施',
            '导线风摆整治措施'
        ]
    },
    '鸟巢故障': {
        '原因': [
            '鸟类在线路杆塔上筑巢',
            '鸟窝被大风吹散落在导线上',
            '树枝或金属丝落在导线上'
        ],
        '现象': [
            '接地',
            '跳闸',
            '线路故障',
            '短路接地'
        ],
        '后果': [
            '线路跳闸',
            '设备损坏'
        ],
        '措施': [
            '防鸟刺',
            '惊鸟牌',
            '防鸟封堵箱',
            '风动驱鸟器',
            '电子驱鸟器',
            '超声波驱鸟器',
            '基于超级电容驱动的组合式驱鸟器'
        ]
    },
    '鸟粪故障': {
        '原因': [
            '鸟粪污染绝缘子串',
            '鸟粪在绝缘子串边沿形成导电通道',
            '鸟粪电导率高',
            '空气潮湿大雾'
        ],
        '现象': [
            '绝缘子闪络',
            '单相接地',
            '空气间隙击穿'
        ],
        '后果': [
            '线路跳闸',
            '设备损坏'
        ],
        '措施': [
            '防鸟刺',
            '惊鸟牌',
            '防鸟封堵箱',
            '绝缘防鸟挡板',
            '横担封堵板',
            '防鸟粪均压环'
        ]
    },
    '复合绝缘子鸟啄损伤': {
        '原因': [
            '鸟类啄击复合绝缘子'
        ],
        '现象': [
            '绝缘子损伤',
            '绝缘性能下降'
        ],
        '后果': [
            '设备损坏',
            '潜在故障风险'
        ],
        '措施': [
            '绝缘防护装置',
            '复合绝缘子保护套'
        ]
    },
    '鸟类飞行短路': {
        '原因': [
            '体形较大的鸟类空中争斗',
            '鸟类在导线间穿越飞行'
        ],
        '现象': [
            '相间短路',
            '单相接地'
        ],
        '后果': [
            '线路跳闸'
        ],
        '措施': [
            '防鸟刺',
            '惊鸟牌',
            '风动驱鸟器',
            '电子驱鸟器'
        ]
    },
    '蛇类捕食短路': {
        '原因': [
            '蛇向塔顶攀爬捕食鸟类',
            '蛇体形较长短接绝缘子'
        ],
        '现象': [
            '绝缘子闪络'
        ],
        '后果': [
            '线路跳闸'
        ],
        '措施': [
            '防鸟封堵箱',
            '绝缘防护装置',
            '复合绝缘子保护套'
        ]
    },
    '绝缘子积污': {
        '原因': [
            '绝缘表面附着污秽物',
            '可溶物质溶于水',
            '形成导电膜'
        ],
        '现象': [
            '绝缘水平降低',
            '强烈放电',
            '闪络'
        ],
        '后果': [
            '线路跳闸',
            '设备损坏'
        ],
        '措施': [
            '采用复合绝缘子防污闪技术',
            '使用防污闪涂料',
            '积极进行绝缘子清扫'
        ]
    },
    '绝缘子劣化': {
        '原因': [
            '瓷玻璃绝缘子长期运行',
            '绝缘子随时间增长使绝缘性能下降',
            '丧失机械支撑能力'
        ],
        '现象': [
            '零值绝缘子',
            '低值绝缘子',
            '绝缘性能下降'
        ],
        '后果': [
            '闪络故障',
            '设备损坏'
        ],
        '措施': [
            '采用复合绝缘子防污闪技术',
            '定期检测零值低值绝缘子',
            '更换劣化绝缘子',
            '积极进行绝缘子清扫'
        ]
    },
    '复合绝缘子憎水性丧失': {
        '原因': [
            '硅橡胶老化',
            '污染物积累',
            '复合绝缘子中低分子流失',
            '环境潮湿'
        ],
        '现象': [
            '憎水性下降',
            '闪络'
        ],
        '后果': [
            '线路跳闸',
            '设备损坏'
        ],
        '措施': [
            '更换老化复合绝缘子',
            '使用防污闪涂料',
            '积极进行绝缘子清扫'
        ]
    },
    '覆冰绝缘子闪络': {
        '原因': [
            '绝缘子覆冰',
            '覆雪'
        ],
        '现象': [
            '闪络'
        ],
        '后果': [
            '线路跳闸',
            '设备损坏'
        ],
        '措施': [
            '增加绝缘子串长度',
            '采用大盘径绝缘子插花方式',
            '采用防冰涂料或防污闪涂料',
            '保持绝缘子和塔头上部的清洁'
        ]
    }
}

# 安全风险
SAFETY_RISKS = [
    '触电风险',
    '高空作业风险',
    '设备损坏风险',
    '火灾风险',
    '停电损失风险',
    '人员伤亡风险',
    '环境影响风险',
    '结构失稳风险',
    '电击伤害风险'
]

# 应急资源
EMERGENCY_RESOURCES = [
    '抢修队伍',
    '备用设备',
    '绝缘工具',
    '起重机械',
    '照明设备',
    '通信设备',
    '交通工具',
    '安全防护用品',
    '监测设备',
    '应急电源',
    '备品备件',
    '金具器材'
]

def create_nodes():
    """创建节点数据"""
    nodes = []

    # 添加电气设备节点
    nodes.append({
        '节点名称': EQUIPMENT,
        '描述': '输电线路是电力系统的重要组成部分，承担电能输送任务',
        '图像': '',
        '权重': 10,
        '度': 0,
        '数值': 1,
        '横坐标': None,
        '纵坐标': None,
        '填充颜色': None,
        '文字颜色': None
    })

    # 添加故障类型节点
    for fault_name, fault_data in FAULTS.items():
        desc = f"输电线路{fault_name}故障"
        if fault_data.get('原因'):
            desc += f"，主要原因：{fault_data['原因'][0]}"
        nodes.append({
            '节点名称': fault_name,
            '描述': desc,
            '图像': '',
            '权重': 5,
            '度': 0,
            '数值': 2,
            '横坐标': None,
            '纵坐标': None,
            '填充颜色': None,
            '文字颜色': None
        })

    # 添加故障原因节点
    reason_set = set()
    for fault_data in FAULTS.values():
        for reason in fault_data.get('原因', []):
            reason_set.add(reason)
    for reason in reason_set:
        nodes.append({
            '节点名称': reason,
            '描述': f"导致输电线路故障的原因：{reason}",
            '图像': '',
            '权重': 3,
            '度': 0,
            '数值': 3,
            '横坐标': None,
            '纵坐标': None,
            '填充颜色': None,
            '文字颜色': None
        })

    # 添加故障现象节点
    phenomenon_set = set()
    for fault_data in FAULTS.values():
        for phenomenon in fault_data.get('现象', []):
            phenomenon_set.add(phenomenon)
    for phenomenon in phenomenon_set:
        nodes.append({
            '节点名称': phenomenon,
            '描述': f"输电线路故障时表现出的现象：{phenomenon}",
            '图像': '',
            '权重': 2,
            '度': 0,
            '数值': 4,
            '横坐标': None,
            '纵坐标': None,
            '填充颜色': None,
            '文字颜色': None
        })

    # 添加故障后果节点
    consequence_set = set()
    for fault_data in FAULTS.values():
        for consequence in fault_data.get('后果', []):
            consequence_set.add(consequence)
    for consequence in consequence_set:
        nodes.append({
            '节点名称': consequence,
            '描述': f"输电线路故障造成的后果：{consequence}",
            '图像': '',
            '权重': 4,
            '度': 0,
            '数值': 5,
            '横坐标': None,
            '纵坐标': None,
            '填充颜色': None,
            '文字颜色': None
        })

    # 添加应对措施节点
    measure_set = set()
    for fault_data in FAULTS.values():
        for measure in fault_data.get('措施', []):
            measure_set.add(measure)
    for measure in measure_set:
        nodes.append({
            '节点名称': measure,
            '描述': f"预防和处理输电线路故障的措施：{measure}",
            '图像': '',
            '权重': 3,
            '度': 0,
            '数值': 6,
            '横坐标': None,
            '纵坐标': None,
            '填充颜色': None,
            '文字颜色': None
        })

    # 添加安全风险节点
    for risk in SAFETY_RISKS:
        nodes.append({
            '节点名称': risk,
            '描述': f"输电线路运行中存在的安全风险：{risk}",
            '图像': '',
            '权重': 2,
            '度': 0,
            '数值': 7,
            '横坐标': None,
            '纵坐标': None,
            '填充颜色': None,
            '文字颜色': None
        })

    # 添加应急资源节点
    for resource in EMERGENCY_RESOURCES:
        nodes.append({
            '节点名称': resource,
            '描述': f"处理输电线路故障所需的应急资源：{resource}",
            '图像': '',
            '权重': 1,
            '度': 0,
            '数值': 8,
            '横坐标': None,
            '纵坐标': None,
            '填充颜色': None,
            '文字颜色': None
        })

    return pd.DataFrame(nodes)

def create_relations():
    """创建关系数据"""
    relations = []

    # 电气设备 -> 故障类型 (发生)
    for fault_name in FAULTS.keys():
        relations.append({
            '源节点': EQUIPMENT,
            '目标节点': fault_name,
            '关系名称': '发生'
        })

    # 故障类型 -> 故障原因 (起因于)
    for fault_name, fault_data in FAULTS.items():
        for reason in fault_data.get('原因', []):
            relations.append({
                '源节点': fault_name,
                '目标节点': reason,
                '关系名称': '起因于'
            })

    # 故障类型 -> 故障现象 (表现为)
    for fault_name, fault_data in FAULTS.items():
        for phenomenon in fault_data.get('现象', []):
            relations.append({
                '源节点': fault_name,
                '目标节点': phenomenon,
                '关系名称': '表现为'
            })

    # 故障类型 -> 故障后果 (导致)
    for fault_name, fault_data in FAULTS.items():
        for consequence in fault_data.get('后果', []):
            relations.append({
                '源节点': fault_name,
                '目标节点': consequence,
                '关系名称': '导致'
            })

    # 故障类型 -> 应对措施 (处置)
    for fault_name, fault_data in FAULTS.items():
        for measure in fault_data.get('措施', []):
            relations.append({
                '源节点': fault_name,
                '目标节点': measure,
                '关系名称': '处置'
            })

    # 故障原因 -> 应对措施 (针对性)
    # 查找每个原因对应的措施
    reason_to_measure = {}
    for fault_name, fault_data in FAULTS.items():
        for measure in fault_data.get('措施', []):
            for reason in fault_data.get('原因', []):
                if reason not in reason_to_measure:
                    reason_to_measure[reason] = set()
                reason_to_measure[reason].add(measure)

    for reason, measures in reason_to_measure.items():
        for measure in measures:
            relations.append({
                '源节点': reason,
                '目标节点': measure,
                '关系名称': '针对性'
            })

    # 应对措施 -> 安全风险 (存在)
    # 某些措施存在特定风险
    measure_risk_mapping = {
        '对杆塔紧固件进行有计划的二次复紧': '高空作业风险',
        '更换老化复合绝缘子': '高空作业风险',
        '更换劣化绝缘子': '高空作业风险',
        '积极进行绝缘子清扫': '高空作业风险',
        '安装线路避雷器': '高空作业风险',
        '采用双绝缘子串加装支撑管改造': '高空作业风险',
        '规范跳线安装和改造措施': '高空作业风险',
        '绝缘工具': '触电风险',
        '起重机械': '设备损坏风险',
        '使用防污闪涂料': '火灾风险',
        '盗窃': '人员伤亡风险',
        '火灾风险': '火灾风险'
    }

    for measure, risk in measure_risk_mapping.items():
        relations.append({
            '源节点': measure,
            '目标节点': risk,
            '关系名称': '存在'
        })

    # 应对措施 -> 应急资源 (需要)
    measure_resource_mapping = {
        '对杆塔紧固件进行有计划的二次复紧': ['抢修队伍', '绝缘工具', '安全防护用品'],
        '更换老化复合绝缘子': ['抢修队伍', '备用设备', '绝缘工具', '安全防护用品'],
        '更换劣化绝缘子': ['抢修队伍', '备用设备', '绝缘工具', '安全防护用品'],
        '积极进行绝缘子清扫': ['抢修队伍', '绝缘工具', '安全防护用品'],
        '安装线路避雷器': ['抢修队伍', '备用设备', '绝缘工具', '安全防护用品'],
        '采用双绝缘子串加装支撑管改造': ['抢修队伍', '起重机械', '绝缘工具', '安全防护用品'],
        '规范跳线安装和改造措施': ['抢修队伍', '绝缘工具', '安全防护用品'],
        '合理选用金具': ['备品备件', '金具器材'],
        '采用高强度、大容量、自重轻的导线': ['备用设备', '交通工具'],
        '起重机械': ['起重机械'],
        '安装移动视频监控及预警装置': ['监测设备', '通信设备'],
        '加强与政府职能部门联系': ['通信设备'],
        '根据季节特点和山火发生规律加强防山火预测及预警': ['监测设备', '通信设备']
    }

    for measure, resources in measure_resource_mapping.items():
        for resource in resources:
            relations.append({
                '源节点': measure,
                '目标节点': resource,
                '关系名称': '需要'
            })

    return pd.DataFrame(relations)

def apply_template_style(target_wb, template_wb):
    """应用模板样式到目标文件"""
    template_ws = template_wb.active
    target_ws = target_wb.active

    # 复制列宽
    for i in range(1, template_ws.max_column + 1):
        col_letter = chr(64 + i) if i <= 26 else f"A{chr(64 + i - 26)}"
        target_ws.column_dimensions[col_letter].width = template_ws.column_dimensions[col_letter].width

    # 复制表头样式
    for cell in template_ws[1]:
        col_letter = cell.column_letter
        target_cell = target_ws[f"{col_letter}1"]
        if cell.has_style:
            target_cell.font = copy(cell.font)
            target_cell.fill = copy(cell.fill)
            target_cell.border = copy(cell.border)
            target_cell.alignment = copy(cell.alignment)

    # 设置行高
    target_ws.row_dimensions[1].height = template_ws.row_dimensions[1].height

def main():
    # 创建节点Excel文件
    nodes_df = create_nodes()
    nodes_file = r'C:\Users\StackRat\Desktop\docker\dataset\xls\输电线路-节点.xlsx'

    # 加载模板
    template_nodes = load_workbook(r'C:\Users\StackRat\Desktop\docker\dataset\xls\NRD Studio Excel模板文件\节点_nodes.xlsx')

    # 保存数据到新文件
    nodes_df.to_excel(nodes_file, index=False)

    # 应用模板样式
    nodes_wb = load_workbook(nodes_file)
    apply_template_style(nodes_wb, template_nodes)
    nodes_wb.save(nodes_file)

    print(f"节点文件已创建: {nodes_file}")
    print(f"节点数量: {len(nodes_df)}")

    # 创建关系Excel文件
    relations_df = create_relations()
    relations_file = r'C:\Users\StackRat\Desktop\docker\dataset\xls\输电线路-关系.xlsx'

    # 加载模板
    template_links = load_workbook(r'C:\Users\StackRat\Desktop\docker\dataset\xls\NRD Studio Excel模板文件\关系_links.xlsx')

    # 保存数据到新文件
    relations_df.to_excel(relations_file, index=False)

    # 应用模板样式
    relations_wb = load_workbook(relations_file)
    apply_template_style(relations_wb, template_links)
    relations_wb.save(relations_file)

    print(f"关系文件已创建: {relations_file}")
    print(f"关系数量: {len(relations_df)}")

    # 打印统计信息
    print("\n=== 数据统计 ===")
    print(f"故障类型数量: {len(FAULTS)}")
    print(f"安全风险数量: {len(SAFETY_RISKS)}")
    print(f"应急资源数量: {len(EMERGENCY_RESOURCES)}")

    # 统计关系类型
    print("\n=== 关系统计 ===")
    relation_counts = relations_df['关系名称'].value_counts()
    for rel_type, count in relation_counts.items():
        print(f"{rel_type}: {count}")

if __name__ == '__main__':
    main()
