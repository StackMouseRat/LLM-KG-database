# -*- coding: utf-8 -*-
from openpyxl import load_workbook

# 验证节点文件
print("=== 验证节点.xlsx ===")
wb = load_workbook(r'C:\Users\StackRat\Desktop\docker\dataset\xls\节点.xlsx')
ws = wb.active
for i, row in enumerate(ws.iter_rows(values_only=True), 1):
    print(f"Row {i}: {row}")

print("\n=== 验证关系.xlsx ===")
wb = load_workbook(r'C:\Users\StackRat\Desktop\docker\dataset\xls\关系.xlsx')
ws = wb.active
for i, row in enumerate(ws.iter_rows(values_only=True), 1):
    print(f"Row {i}: {row}")
