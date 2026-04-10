from __future__ import annotations

import csv
import re
from pathlib import Path
from typing import Dict, List, Tuple

from openpyxl import load_workbook


SEP = "<<<CHUNK>>>"
TXT_COMMON = "\u901a\u7528"
TXT_TEMPLATE = "template_clause"
TXT_SLOT_FACT = "slot_fact"
SLOT_FAULT_OBJECT = "fault_object"

TXT_CAUSE = "\u8d77\u56e0\u4e8e"
TXT_SHOW = "\u8868\u73b0"
SUFFIX_CAUSE = "-\u6545\u969c\u539f\u56e0"
SUFFIX_PHENO = "-\u6545\u969c\u73b0\u8c61"

GENERIC_OBJECT = (
    "\u9ad8\u538b\u65ad\u8def\u5668\u672c\u4f53\u53ca\u5176\u706d\u5f27\u5ba4\u3001"
    "\u7edd\u7f18\u652f\u6491\u7ed3\u6784\u3001\u64cd\u52a8\u673a\u6784\u3001"
    "\u5bfc\u7535\u56de\u8def\u3001\u63a7\u5236\u4e0e\u6d4b\u91cf\u4e8c\u6b21\u56de\u8def"
    "\u76f8\u5173\u88c5\u7f6e"
)


def find_single(pattern: str, root: Path) -> Path:
    matches = list(root.glob(pattern))
    if not matches:
        raise FileNotFoundError(f"Pattern not found: {pattern}")
    return sorted(matches, key=lambda p: len(str(p)))[0]


def trim_cn_period(text: str) -> str:
    t = (text or "").strip()
    return t[:-1] if t.endswith("\u3002") else t


def load_node_desc_map(nodes_xlsx: Path) -> Dict[str, str]:
    wb = load_workbook(nodes_xlsx, read_only=True, data_only=True)
    ws = wb[wb.sheetnames[0]]
    rows = ws.iter_rows(values_only=True)
    header = next(rows, None)
    if not header:
        return {}

    idx_name = -1
    idx_desc = -1
    for i, h in enumerate(header):
        if h == "name":
            idx_name = i
        elif h == "desc":
            idx_desc = i
    if idx_name < 0 or idx_desc < 0:
        return {}

    out: Dict[str, str] = {}
    for row in rows:
        if not row:
            continue
        name = str(row[idx_name]).strip() if idx_name < len(row) and row[idx_name] is not None else ""
        desc = str(row[idx_desc]).strip() if idx_desc < len(row) and row[idx_desc] is not None else ""
        if name and desc:
            out[name] = desc
    return out


def normalize_label(text: str) -> str:
    t = (text or "").strip()
    t = t.replace(SUFFIX_CAUSE, "").replace(SUFFIX_PHENO, "").replace("-\u6545\u969c\u540e\u679c", "")
    t = re.sub(r"\s+", "", t)
    return t


def normalize_relation_sentence(text: str, node_desc_map: Dict[str, str]) -> str:
    content = (text or "").strip()
    if not content:
        return content

    # X起因于Y-故障原因
    if TXT_CAUSE in content and SUFFIX_CAUSE in content:
        left, right = content.split(TXT_CAUSE, 1)
        left = left.strip()
        right_clean = right.strip()
        right_base = right_clean.replace(SUFFIX_CAUSE, "").strip()
        cause_node = f"{right_base}{SUFFIX_CAUSE}"
        desc = trim_cn_period(node_desc_map.get(cause_node, ""))
        if desc:
            return f"{left}\u7684\u6545\u969c\u539f\u56e0\u4e3a\uff1a{desc}\u3002"

    # X-故障原因表现Y-故障现象
    if TXT_SHOW in content and SUFFIX_CAUSE in content and SUFFIX_PHENO in content:
        left, right = content.split(TXT_SHOW, 1)
        left_base = left.replace(SUFFIX_CAUSE, "").strip()
        right_base = right.replace(SUFFIX_PHENO, "").strip()
        pheno_node = f"{right_base}{SUFFIX_PHENO}"
        desc = trim_cn_period(node_desc_map.get(pheno_node, ""))
        if desc:
            return f"{left_base}\u7684\u6545\u969c\u73b0\u8c61\u4e3a\uff1a{desc}\u3002"

    # Self-reference fallback
    c = trim_cn_period(content)
    if TXT_CAUSE in c:
        left, right = c.split(TXT_CAUSE, 1)
        if normalize_label(left) == normalize_label(right):
            desc = trim_cn_period(node_desc_map.get(normalize_label(left), ""))
            if desc:
                return f"{normalize_label(left)}\u7684\u6545\u969c\u8bf1\u56e0\u53ef\u53c2\u8003\uff1a{desc}\u3002"
    if "\u7684\u6545\u969c\u73b0\u8c61\u8868\u73b0\u4e3a" in c:
        left, right = c.split("\u7684\u6545\u969c\u73b0\u8c61\u8868\u73b0\u4e3a", 1)
        if normalize_label(left) == normalize_label(right):
            desc = trim_cn_period(node_desc_map.get(normalize_label(left), ""))
            if desc:
                return f"{normalize_label(left)}\u7684\u5178\u578b\u6545\u969c\u73b0\u8c61\u53ef\u53c2\u8003\uff1a{desc}\u3002"

    return content


def enhance_with_desc(text: str, node_desc_map: Dict[str, str], max_terms: int) -> str:
    base = trim_cn_period(text)
    if not base:
        return base

    matches: List[Tuple[str, str]] = []
    for name in sorted(node_desc_map.keys(), key=len, reverse=True):
        if name in base:
            desc = trim_cn_period(node_desc_map[name])
            if desc and desc not in base:
                matches.append((name, desc))
        if len(matches) >= max_terms:
            break

    if not matches:
        return base

    details = "\uff1b".join(f"{name}\uff1a{desc}" for name, desc in matches)
    return f"{base} \u76f8\u5173\u5b9a\u4e49\uff1a{details}\u3002"


def write_csv(rows: List[dict], out_csv: Path) -> None:
    fields = [
        "chunk_id",
        "doc_id",
        "chunk_type",
        "asset_type",
        "fault_type",
        "slot_key",
        "content",
        "keywords",
        "source_ref",
        "priority",
        "version",
    ]
    with out_csv.open("w", encoding="utf-8-sig", newline="") as f:
        w = csv.DictWriter(f, fieldnames=fields)
        w.writeheader()
        for r in rows:
            w.writerow({k: r.get(k, "") for k in fields})


def write_txt_for_fastgpt(rows: List[dict], out_txt: Path) -> None:
    with out_txt.open("w", encoding="utf-8-sig") as f:
        for i, r in enumerate(rows):
            if i:
                f.write("\n")
            lines = [
                f"chunk_id: {r.get('chunk_id', '')}",
                f"doc_id: {r.get('doc_id', '')}",
                f"chunk_type: {r.get('chunk_type', '')}",
                f"asset_type: {r.get('asset_type', '')}",
                f"fault_type: {r.get('fault_type', '')}",
                f"slot_key: {r.get('slot_key', '')}",
                f"keywords: {r.get('keywords', '')}",
                f"source_ref: {r.get('source_ref', '')}",
                f"priority: {r.get('priority', '')}",
                f"version: {r.get('version', '')}",
                "content:",
                (r.get("content", "") or "").strip(),
                SEP,
            ]
            f.write("\n".join(lines))


def main() -> None:
    root = Path(__file__).resolve().parent
    kb_chunks_csv = find_single("**/csv/kb_chunks.csv", root)
    source_dir = find_single("xls/**/high_voltage_breaker_1to8_import4", root)
    nodes_xlsx = find_single(str(source_dir.relative_to(root) / "*_nodes.xlsx"), root)
    node_desc_map = load_node_desc_map(nodes_xlsx)

    with kb_chunks_csv.open("r", encoding="utf-8-sig", newline="") as f:
        rows = list(csv.DictReader(f))

    fixed_rows: List[dict] = []
    relation_norm_count = 0
    desc_enhance_count = 0

    for row in rows:
        r = dict(row)
        if r.get("chunk_type") == TXT_SLOT_FACT:
            original = (r.get("content", "") or "").strip()

            if r.get("slot_key") == SLOT_FAULT_OBJECT:
                fault_type = (r.get("fault_type", "") or "").strip()
                if fault_type and fault_type != TXT_COMMON:
                    original = f"{fault_type}\u53ef\u80fd\u53d1\u751f\u4e8e{GENERIC_OBJECT}\u3002"
                    r["priority"] = "4"

            normalized = normalize_relation_sentence(original, node_desc_map)
            if normalized != original:
                relation_norm_count += 1

            max_terms = 1 if r.get("slot_key") == SLOT_FAULT_OBJECT else 2
            enhanced = enhance_with_desc(normalized, node_desc_map, max_terms=max_terms)
            if enhanced != normalized:
                desc_enhance_count += 1
            r["content"] = enhanced

        fixed_rows.append(r)

    # Dedup by semantic key.
    seen = set()
    final_rows: List[dict] = []
    for r in fixed_rows:
        key = (
            r.get("chunk_type", ""),
            r.get("asset_type", ""),
            r.get("fault_type", ""),
            r.get("slot_key", ""),
            r.get("content", ""),
        )
        if key in seen:
            continue
        seen.add(key)
        final_rows.append(r)

    out_csv = kb_chunks_csv.parent / "kb_chunks_corrected_v5_name_desc.csv"
    write_csv(final_rows, out_csv)

    retrieval_rows = [r for r in final_rows if r.get("chunk_type") != TXT_TEMPLATE]
    out_txt = kb_chunks_csv.parent / "kb_chunks_fastgpt_upload_retrieval_only_corrected_v5_name_desc.txt"
    write_txt_for_fastgpt(retrieval_rows, out_txt)

    print("kb_chunks_csv:", kb_chunks_csv)
    print("nodes_xlsx:", nodes_xlsx)
    print("node_desc_count:", len(node_desc_map))
    print("original_rows:", len(rows))
    print("final_rows:", len(final_rows))
    print("retrieval_rows:", len(retrieval_rows))
    print("relation_norm_count:", relation_norm_count)
    print("desc_enhance_count:", desc_enhance_count)
    print("out_csv:", out_csv)
    print("out_txt:", out_txt)


if __name__ == "__main__":
    main()
